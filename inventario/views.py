from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q
from .models import Equipo
from .forms import EquipoForm
from bitacora.models import Auditoria, RegistroBitacora
from django.http import HttpResponse
import openpyxl
from io import BytesIO
from django.template.loader import get_template
from django.template.loader import render_to_string
from weasyprint import HTML


def es_admin(user):
    return user.is_staff

@login_required
def lista_equipos(request):
    query = request.GET.get('q')
    equipos = Equipo.objects.all()
    if query:
        equipos = equipos.filter(
            Q(nombre__icontains=query) |
            Q(marca__icontains=query) |
            Q(modelo__icontains=query) |
            Q(serial__icontains=query)
        ).distinct()
    return render(request, 'inventario/lista_equipos.html', {'equipos': equipos})

@login_required
def detalle_equipo(request, pk):
    equipo = get_object_or_404(Equipo, pk=pk)
    registros = RegistroBitacora.objects.filter(equipo=equipo).order_by('-fecha')
    return render(request, 'inventario/detalle_equipo.html', {'equipo': equipo, 'registros': registros})

@login_required
@user_passes_test(es_admin)
def agregar_equipo(request):
    if request.method == 'POST':
        form = EquipoForm(request.POST)
        if form.is_valid():
            equipo = form.save()
            Auditoria.objects.create(usuario=request.user, accion=f"Se creó el equipo: {equipo.nombre}")
            return redirect('lista_equipos')
    else:
        form = EquipoForm()
    return render(request, 'inventario/form_equipo.html', {'form': form, 'titulo': 'Agregar Equipo'})

@login_required
@user_passes_test(es_admin)
def modificar_equipo(request, pk):
    equipo = get_object_or_404(Equipo, pk=pk)
    if request.method == 'POST':
        form = EquipoForm(request.POST, instance=equipo)
        if form.is_valid():
            equipo = form.save()
            Auditoria.objects.create(usuario=request.user, accion=f"Se modificó el equipo: {equipo.nombre}")
            return redirect('detalle_equipo', pk=equipo.pk)
    else:
        form = EquipoForm(instance=equipo)
    return render(request, 'inventario/form_equipo.html', {'form': form, 'titulo': 'Modificar Equipo'})

@login_required
@user_passes_test(es_admin)
def eliminar_equipo(request, pk):
    equipo = get_object_or_404(Equipo, pk=pk)
    if request.method == 'POST':
        Auditoria.objects.create(usuario=request.user, accion=f"Se eliminó el equipo: {equipo.nombre}")
        equipo.delete()
        return redirect('lista_equipos')
    return render(request, 'inventario/confirmar_eliminar.html', {'equipo': equipo})

@login_required
def exportar_equipos_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario_equipos.xlsx"'
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Inventario IT"
    
    columns = ['Nombre', 'Tipo', 'Marca', 'Modelo', 'Serial', 'Ubicación', 'Estado', 'Fecha Adquisición']
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for equipo in Equipo.objects.all():
        row_num += 1
        row = [
            equipo.nombre,
            equipo.get_tipo_display(),
            equipo.marca,
            equipo.modelo,
            equipo.serial,
            equipo.ubicacion,
            equipo.get_estado_display(),
            equipo.fecha_adquisicion.strftime('%Y-%m-%d'),
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    
    workbook.save(response)
    return response

@login_required
def exportar_equipos_pdf(request):
    equipos = Equipo.objects.all()
    html_string = render_to_string('inventario/reporte_pdf.html', {'equipos': equipos})
    html = HTML(string=html_string)
    pdf = html.write_pdf()
    
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'filename="inventario_equipos.pdf"'
    return response
